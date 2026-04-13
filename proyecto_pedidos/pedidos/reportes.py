from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from datetime import datetime

# ReportLab: Generación de PDFs
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

# Openpyxl: Excel
from openpyxl import Workbook

# Modelos
from .models import Pedido, DetallePedido
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from proyecto_pedidos.clientes.models import Cliente
from proyecto_pedidos.productos.models import Producto

from .models import DetallePedido, Pedido


# aplica estilo visual a tablas pdf
def _estilo_tabla(table):
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))


# agrega encabezado comun para reportes pdf
def _encabezado_pdf(elements, styles, titulo, total_label, total):
    fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
    elements.append(Paragraph('SISTEMA DE GESTION', styles['Normal']))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph(titulo, styles['Title']))
    elements.append(Spacer(1, 10))
    elements.append(Paragraph(f'Fecha de generacion: {fecha}', styles['Normal']))
    if total_label:
        elements.append(Paragraph(f'{total_label}: {total}', styles['Normal']))
    elements.append(Spacer(1, 20))


@login_required
# exporta pedidos en formato pdf
def exportar_pedidos_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="pedidos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    pedidos = Pedido.objects.all()

    _encabezado_pdf(elements, styles, 'REPORTE DE PEDIDOS', 'Total de pedidos', pedidos.count())

    data = [['ID', 'Cliente', 'Fecha', 'Estado']]
    for pedido in pedidos:
        data.append([pedido.id, pedido.cliente.nombre, str(pedido.fecha), pedido.estado])

    table = Table(data, colWidths=[50, 170, 110, 110])
    _estilo_tabla(table)
    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph('Documento generado automaticamente por el sistema', styles['Italic']))

    doc.build(elements)
    return response


@login_required
# exporta clientes en formato pdf
def exportar_clientes_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="clientes.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    clientes = Cliente.objects.all()

    _encabezado_pdf(elements, styles, 'REPORTE DE CLIENTES', 'Total de clientes', clientes.count())

    data = [['ID', 'Nombre', 'Correo', 'Telefono']]
    for cliente in clientes:
        data.append([cliente.id, cliente.nombre, cliente.correo, cliente.telefono])

    table = Table(data, colWidths=[45, 145, 220, 100])
    _estilo_tabla(table)
    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph('Documento generado automaticamente por el sistema', styles['Italic']))

    doc.build(elements)
    return response


@login_required
# exporta productos en formato pdf
def exportar_productos_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="productos.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    productos = Producto.objects.all()

    _encabezado_pdf(elements, styles, 'REPORTE DE PRODUCTOS', 'Total de productos', productos.count())

    data = [['ID', 'Nombre', 'Precio', 'Stock']]
    for producto in productos:
        data.append([producto.id, producto.nombre, str(producto.precio), producto.stock])

    table = Table(data, colWidths=[45, 185, 90, 70])
    _estilo_tabla(table)
    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph('Documento generado automaticamente por el sistema', styles['Italic']))

    doc.build(elements)
    return response


@login_required
# exporta detalles en formato pdf
def exportar_detalles_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="detalles.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    detalles = DetallePedido.objects.all()

    _encabezado_pdf(elements, styles, 'REPORTE DE DETALLES', None, None)

    data = [['Pedido', 'Cliente', 'Producto', 'Cant', 'Subtotal']]
    for detalle in detalles:
        data.append([
            detalle.pedido.id,
            detalle.pedido.cliente.nombre,
            detalle.producto.nombre,
            detalle.cantidad,
            str(detalle.subtotal),
        ])

    table = Table(data, colWidths=[45, 120, 190, 55, 80])
    _estilo_tabla(table)
    elements.append(table)
    elements.append(Spacer(1, 30))
    elements.append(Paragraph('Documento generado automaticamente por el sistema', styles['Italic']))

    doc.build(elements)
    return response


@login_required
# exporta pedidos en formato excel
def exportar_pedidos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos'
    ws.append(['ID', 'Cliente', 'Fecha', 'Estado'])
    for pedido in Pedido.objects.all():
        ws.append([pedido.id, pedido.cliente.nombre, str(pedido.fecha), pedido.estado])
    wb.save(response)
    return response


@login_required
# exporta clientes en formato excel
def exportar_clientes_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=clientes.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Clientes'
    ws.append(['ID', 'Nombre', 'Correo', 'Telefono', 'Direccion'])
    for cliente in Cliente.objects.all():
        ws.append([cliente.id, cliente.nombre, cliente.correo, cliente.telefono, cliente.direccion])
    wb.save(response)
    return response


@login_required
# exporta productos en formato excel
def exportar_productos_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'
    ws.append(['ID', 'Nombre', 'Precio', 'Stock'])
    for producto in Producto.objects.all():
        ws.append([producto.id, producto.nombre, float(producto.precio), producto.stock])
    wb.save(response)
    return response


@login_required
# exporta detalles en formato excel
def exportar_detalles_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=detalles.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Detalles'
    ws.append(['ID Detalle', 'Pedido', 'Cliente', 'Producto', 'Cantidad', 'Subtotal'])
    for detalle in DetallePedido.objects.all():
        ws.append([
            detalle.id,
            detalle.pedido.id,
            detalle.pedido.cliente.nombre,
            detalle.producto.nombre,
            detalle.cantidad,
            float(detalle.subtotal),
        ])
    wb.save(response)
    return response


@login_required
# exporta todas las hojas en un solo excel
def exportar_todo_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=reportes.xlsx'

    wb = Workbook()
    ws_pedidos = wb.active
    ws_pedidos.title = 'Pedidos'
    ws_pedidos.append(['ID', 'Cliente', 'Fecha', 'Estado'])
    for pedido in Pedido.objects.all():
        ws_pedidos.append([pedido.id, pedido.cliente.nombre, str(pedido.fecha), pedido.estado])

    ws_clientes = wb.create_sheet(title='Clientes')
    ws_clientes.append(['ID', 'Nombre', 'Correo', 'Telefono', 'Direccion'])
    for cliente in Cliente.objects.all():
        ws_clientes.append([cliente.id, cliente.nombre, cliente.correo, cliente.telefono, cliente.direccion])

    ws_productos = wb.create_sheet(title='Productos')
    ws_productos.append(['ID', 'Nombre', 'Precio', 'Stock'])
    for producto in Producto.objects.all():
        ws_productos.append([producto.id, producto.nombre, float(producto.precio), producto.stock])

    wb.save(response)
    return response
    wb = Workbook()
